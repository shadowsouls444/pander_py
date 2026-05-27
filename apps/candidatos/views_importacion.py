"""
apps/candidatos/views_importacion.py
Importación masiva desde Excel con:
  - Validación completa antes de importar (si hay errores → no importa nada)
  - Log de validación por fila con descripción del error
  - Relaciones por nombre/tipo+número de documento (no por ID)
  - Envío automático de emails al importar postulaciones (via trigger)
"""
import io
import hashlib
import secrets
import string
from datetime import timedelta

from django.utils import timezone
from django.conf import settings
from django.core.mail import send_mail
from django.db import transaction
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser

try:
    import openpyxl
except ImportError:
    openpyxl = None


def _hash(raw: str) -> str:
    return hashlib.sha256(raw.encode()).hexdigest()


def _gen_pwd(n: int = 12) -> str:
    chars = string.ascii_letters + string.digits + "@#$%&*"
    pwd = [secrets.choice(string.ascii_uppercase),
           secrets.choice(string.ascii_lowercase),
           secrets.choice(string.digits),
           secrets.choice("@#$%&*")]
    pwd += [secrets.choice(chars) for _ in range(n - 4)]
    secrets.SystemRandom().shuffle(pwd)
    return "".join(pwd)


def _gen_login(primer_nombre: str, primer_apellido: str) -> str:
    from apps.acceso.models import Usuario
    base = (primer_nombre[0] + primer_apellido).lower()
    base = "".join(c for c in base if c.isalnum())
    n = Usuario.objects.filter(login__startswith=base).count()
    return f"{base}{n + 1:03d}"


def _leer_excel(archivo):
    """Lee el archivo Excel y retorna lista de dicts por fila."""
    if openpyxl is None:
        raise ImportError("openpyxl no está instalado. Ejecuta: pip install openpyxl")
    wb = openpyxl.load_workbook(archivo, data_only=True)
    ws = wb.active
    headers = [str(c.value or "").strip().lower().replace(" ", "_") for c in ws[1]]
    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in row):
            continue
        rows.append({"_fila": row_idx,
                     **{headers[i]: (str(row[i]).strip() if row[i] is not None else "")
                        for i in range(len(headers))}})
    return rows


def _ok(fila, mensaje="OK"):
    return {"fila": fila, "estado": "OK", "mensaje": mensaje}


def _err(fila, mensaje):
    return {"fila": fila, "estado": "ERROR", "mensaje": mensaje}


# ════════════════════════════════════════════════════════════
# 1 — COMPAÑÍAS
# ════════════════════════════════════════════════════════════

class ImportarCompaniasView(APIView):
    parser_classes = [MultiPartParser]

    COLUMNAS_REQUERIDAS = ["descripcion", "nit"]
    PLANTILLA = [
        ["descripcion", "nit", "objeto_social", "representante_legal",
         "direccion", "telefono", "ind_activa", "ind_evaluacion_vacante"],
        ["Empresa Ejemplo S.A.", "900123456-1", "Comercio al por mayor",
         "Juan Pérez", "Calle 10 #5-20", "6015551234", "TRUE", "FALSE"],
    ]

    def post(self, request):
        return _procesar_importacion(
            request, self.COLUMNAS_REQUERIDAS, self._validar, self._importar)

    def _validar(self, rows, compania_id, **kw):
        from apps.empresa.models import Compania
        log = []
        for r in rows:
            f = r["_fila"]
            if not r.get("descripcion"): log.append(_err(f, "descripcion es obligatorio")); continue
            if not r.get("nit"):         log.append(_err(f, "nit es obligatorio")); continue
            if Compania.objects.filter(nit=r["nit"]).exists():
                log.append(_err(f, f"NIT '{r['nit']}' ya existe")); continue
            log.append(_ok(f, f"Compañía '{r['descripcion']}' lista"))
        return log

    @staticmethod
    def _importar(rows, compania_id, usuario_id, **kw):
        from django.utils import timezone
        from apps.empresa.models import Compania
        from apps.empresa.views import _copiar_analistas_y_superusuarios

        creados = 0
        for r in rows:
            nueva = Compania.objects.create(
                descripcion            = r["descripcion"],
                nit                    = r["nit"],
                objeto_social          = r.get("objeto_social") or None,
                representante_legal    = r.get("representante_legal") or None,
                direccion              = r.get("direccion") or None,
                telefono               = r.get("telefono") or None,
                ind_activa             = r.get("ind_activa", "TRUE").upper() == "TRUE",
                ind_evaluacion_vacante = r.get("ind_evaluacion_vacante", "FALSE").upper() == "TRUE",
                usuario_creacion       = usuario_id,
            )
            # El trigger SQL copia automáticamente la evaluación estándar.
            # Python copia analistas y superusuarios de la compañía 0000.
            _copiar_analistas_y_superusuarios(nueva, uid=usuario_id)
            creados += 1
        return creados


# ════════════════════════════════════════════════════════════
# 2 — ANALISTAS
# ════════════════════════════════════════════════════════════

class ImportarAnalistasView(APIView):
    parser_classes = [MultiPartParser]
    COLUMNAS_REQUERIDAS = ["primer_nombre", "primer_apellido"]

    def post(self, request, compania):
        return _procesar_importacion(
            request, self.COLUMNAS_REQUERIDAS, self._validar, self._importar,
            compania_id=compania)

    def _validar(self, rows, compania_id, **kw):
        from apps.candidatos.models import TipoDocumento
        log = []
        tipos_doc = {t.descripcion.upper(): t.id for t in TipoDocumento.objects.all()}
        for r in rows:
            f = r["_fila"]
            if not r.get("primer_nombre") or not r.get("primer_apellido"):
                log.append(_err(f, "primer_nombre y primer_apellido son obligatorios")); continue
            if r.get("tipo_documento") and r["tipo_documento"].upper() not in tipos_doc:
                log.append(_err(f, f"tipo_documento '{r['tipo_documento']}' no existe")); continue
            log.append(_ok(f, f"Analista '{r['primer_nombre']} {r['primer_apellido']}' listo"))
        return log

    @transaction.atomic
    def _importar(self, rows, compania_id, usuario_id, **kw):
        from apps.acceso.models import Analista
        from apps.candidatos.models import TipoDocumento
        tipos_doc = {t.descripcion.upper(): t for t in TipoDocumento.objects.all()}
        n = Analista.objects.filter(compania=compania_id).count()
        creados = 0
        for r in rows:
            tipo = tipos_doc.get((r.get("tipo_documento") or "").upper())
            n += 1
            Analista.objects.create(
                compania         = _comp(compania_id),
                id_interno       = n,
                tipo_documento   = tipo,
                numero_documento = r.get("numero_documento") or None,
                primer_nombre    = r["primer_nombre"],
                segundo_nombre   = r.get("segundo_nombre") or None,
                primer_apellido  = r["primer_apellido"],
                segundo_apellido = r.get("segundo_apellido") or None,
                telefono         = r.get("telefono") or None,
                cargo            = r.get("cargo") or None,
                usuario_creacion = usuario_id,
            )
            creados += 1
        return creados


# ════════════════════════════════════════════════════════════
# 3 — USUARIOS (relacionar analista por tipo+num documento)
# ════════════════════════════════════════════════════════════

class ImportarUsuariosView(APIView):
    parser_classes = [MultiPartParser]
    COLUMNAS_REQUERIDAS = ["email", "rol"]

    def post(self, request, compania):
        return _procesar_importacion(
            request, self.COLUMNAS_REQUERIDAS, self._validar, self._importar,
            compania_id=compania)

    def _validar(self, rows, compania_id, **kw):
        from apps.acceso.models import Rol, Analista
        from apps.candidatos.models import TipoDocumento
        log = []
        roles = {r.descripcion.upper(): r.id for r in Rol.objects.all()}
        tipos_doc = {t.descripcion.upper(): t.id for t in TipoDocumento.objects.all()}
        for r in rows:
            f = r["_fila"]
            if not r.get("email"): log.append(_err(f, "email es obligatorio")); continue
            if not r.get("rol"):   log.append(_err(f, "rol es obligatorio")); continue
            if r["rol"].upper() not in roles:
                log.append(_err(f, f"Rol '{r['rol']}' no existe")); continue
            # Relación analista por tipo+num documento
            tipo_doc = r.get("tipo_documento_analista", "").upper()
            num_doc  = r.get("numero_documento_analista", "").strip()
            if tipo_doc and num_doc:
                if tipo_doc not in tipos_doc:
                    log.append(_err(f, f"tipo_documento_analista '{tipo_doc}' no existe")); continue
                tid = tipos_doc[tipo_doc]
                if not Analista.objects.filter(
                    compania=compania_id, tipo_documento=tid,
                    numero_documento=num_doc).exists():
                    log.append(_err(f, f"Analista con {tipo_doc} {num_doc} no encontrado")); continue
            log.append(_ok(f, f"Usuario '{r['email']}' listo"))
        return log

    @transaction.atomic
    def _importar(self, rows, compania_id, usuario_id, **kw):
        from apps.acceso.models import Rol, Analista, Usuario
        from apps.candidatos.models import TipoDocumento
        roles = {r.descripcion.upper(): r for r in Rol.objects.all()}
        tipos_doc = {t.descripcion.upper(): t for t in TipoDocumento.objects.all()}
        n = Usuario.objects.filter(compania=compania_id).count()
        creados = 0
        fe = getattr(settings, "FRONTEND_URL", "http://localhost:5173")
        for r in rows:
            analista = None
            tipo_doc = r.get("tipo_documento_analista", "").upper()
            num_doc  = r.get("numero_documento_analista", "").strip()
            if tipo_doc and num_doc:
                tid = tipos_doc[tipo_doc]
                analista = Analista.objects.filter(
                    compania=compania_id, tipo_documento=tid, numero_documento=num_doc).first()
            pnombre = analista.primer_nombre if analista else "Usuario"
            papellido = analista.primer_apellido if analista else ""
            login    = _gen_login(pnombre, papellido)
            pwd_raw  = _gen_pwd()
            n += 1
            u = Usuario.objects.create(
                compania         = _comp(compania_id),
                id_interno       = n,
                analista         = analista,
                rol              = roles[r["rol"].upper()],
                login            = login,
                pwd              = _hash(pwd_raw),
                email            = r["email"],
                ind_super_usuario = r.get("ind_super_usuario","FALSE").upper()=="TRUE",
                ind_activo       = True,
                ind_bloqueo      = False,
                usuario_creacion = usuario_id,
            )
            try:
                send_mail(
                    "Pander RRHH — Tu cuenta ha sido creada",
                    f"Hola {pnombre} {papellido},\n\n"
                    f"Usuario: {login}\nContraseña: {pwd_raw}\n\nPlataforma: {fe}\n\nEquipo Pander RRHH",
                    settings.DEFAULT_FROM_EMAIL, [r["email"]], fail_silently=True)
            except Exception:
                pass
            creados += 1
        return creados


# ════════════════════════════════════════════════════════════
# 4 — UNIDADES ORGANIZACIONALES
# ════════════════════════════════════════════════════════════

class ImportarUnidadesView(APIView):
    parser_classes = [MultiPartParser]
    COLUMNAS_REQUERIDAS = ["descripcion"]

    def post(self, request, compania):
        return _procesar_importacion(
            request, self.COLUMNAS_REQUERIDAS, self._validar, self._importar,
            compania_id=compania)

    def _validar(self, rows, compania_id, **kw):
        log = []
        for r in rows:
            f = r["_fila"]
            if not r.get("descripcion"): log.append(_err(f, "descripcion es obligatorio")); continue
            log.append(_ok(f, f"Unidad '{r['descripcion']}' lista"))
        return log

    @transaction.atomic
    def _importar(self, rows, compania_id, usuario_id, **kw):
        from apps.empresa.models import UnidadOrg
        n = UnidadOrg.objects.filter(compania=compania_id).count()
        creados = 0
        for r in rows:
            n += 1
            UnidadOrg.objects.create(
                compania         = _comp(compania_id),
                id_interno       = n,
                descripcion      = r["descripcion"],
                especialidad     = r.get("especialidad") or None,
                usuario_creacion = usuario_id,
            )
            creados += 1
        return creados


# ════════════════════════════════════════════════════════════
# 5 — VACANTES (relacionar por nombre de unidad)
# ════════════════════════════════════════════════════════════

class ImportarVacantesView(APIView):
    parser_classes = [MultiPartParser]
    COLUMNAS_REQUERIDAS = ["descripcion", "unidad", "estado", "tipo_contrato"]

    def post(self, request, compania):
        return _procesar_importacion(
            request, self.COLUMNAS_REQUERIDAS, self._validar, self._importar,
            compania_id=compania)

    def _validar(self, rows, compania_id, **kw):
        from apps.empresa.models import UnidadOrg
        from apps.vacantes.models import EstadoVacante, TipoContrato
        log = []
        unidades  = {u.descripcion.upper(): u.id for u in UnidadOrg.objects.filter(compania=compania_id)}
        estados   = {e.descripcion.upper(): e.id for e in EstadoVacante.objects.all()}
        contratos = {c.descripcion.upper(): c.id for c in TipoContrato.objects.all()}
        for r in rows:
            f = r["_fila"]
            if not r.get("descripcion"): log.append(_err(f, "descripcion es obligatorio")); continue
            if not r.get("unidad"):      log.append(_err(f, "unidad es obligatorio")); continue
            if r["unidad"].upper() not in unidades:
                log.append(_err(f, f"Unidad '{r['unidad']}' no existe en esta compañía")); continue
            if r.get("estado") and r["estado"].upper() not in estados:
                log.append(_err(f, f"Estado '{r['estado']}' no existe")); continue
            if r.get("tipo_contrato") and r["tipo_contrato"].upper() not in contratos:
                log.append(_err(f, f"Tipo contrato '{r['tipo_contrato']}' no existe")); continue
            log.append(_ok(f, f"Vacante '{r['descripcion'][:40]}' lista"))
        return log

    @transaction.atomic
    def _importar(self, rows, compania_id, usuario_id, **kw):
        from apps.empresa.models import UnidadOrg
        from apps.vacantes.models import EstadoVacante, TipoContrato, Vacante
        unidades  = {u.descripcion.upper(): u for u in UnidadOrg.objects.filter(compania=compania_id)}
        estados   = {e.descripcion.upper(): e for e in EstadoVacante.objects.all()}
        contratos = {c.descripcion.upper(): c for c in TipoContrato.objects.all()}
        n = Vacante.objects.filter(compania=compania_id).count()
        creados = 0
        for r in rows:
            n += 1
            Vacante.objects.create(
                compania         = _comp(compania_id),
                id_interno       = n,
                descripcion      = r["descripcion"],
                unidad           = unidades[r["unidad"].upper()],
                estado           = estados.get(r.get("estado","").upper(), list(estados.values())[0]),
                tipo_contrato    = contratos.get(r.get("tipo_contrato","").upper(), list(contratos.values())[0]),
                anio_experiencia = int(r["anio_experiencia"]) if r.get("anio_experiencia") else None,
                salario_minimo   = float(r["salario_minimo"]) if r.get("salario_minimo") else None,
                salario_maximo   = float(r["salario_maximo"]) if r.get("salario_maximo") else None,
                ind_activa       = r.get("ind_activa","TRUE").upper() == "TRUE",
                ind_publicada    = r.get("ind_publicada","FALSE").upper() == "TRUE",
                usuario_creacion = usuario_id,
            )
            creados += 1
        return creados


# ════════════════════════════════════════════════════════════
# 6 — CANDIDATOS
# ════════════════════════════════════════════════════════════

class ImportarCandidatosView(APIView):
    parser_classes = [MultiPartParser]
    COLUMNAS_REQUERIDAS = ["primer_nombre", "primer_apellido"]

    def post(self, request, compania):
        return _procesar_importacion(
            request, self.COLUMNAS_REQUERIDAS, self._validar, self._importar,
            compania_id=compania)

    def _validar(self, rows, compania_id, **kw):
        from apps.candidatos.models import TipoDocumento
        log = []
        tipos = {t.descripcion.upper(): t.id for t in TipoDocumento.objects.all()}
        for r in rows:
            f = r["_fila"]
            if not r.get("primer_nombre") or not r.get("primer_apellido"):
                log.append(_err(f, "primer_nombre y primer_apellido son obligatorios")); continue
            if r.get("tipo_documento") and r["tipo_documento"].upper() not in tipos:
                log.append(_err(f, f"tipo_documento '{r['tipo_documento']}' no existe")); continue
            log.append(_ok(f, f"Candidato '{r['primer_nombre']} {r['primer_apellido']}' listo"))
        return log

    @transaction.atomic
    def _importar(self, rows, compania_id, usuario_id, **kw):
        from apps.candidatos.models import TipoDocumento, Candidato, DatosCandidato
        tipos = {t.descripcion.upper(): t for t in TipoDocumento.objects.all()}
        n = Candidato.objects.filter(compania=compania_id).count()
        creados = 0
        for r in rows:
            n += 1
            comp = _comp(compania_id)
            cand = Candidato.objects.create(
                compania         = comp,
                id_interno       = n,
                usuario_creacion = usuario_id,
            )
            DatosCandidato.objects.create(
                compania         = comp,
                candidato        = cand,
                tipo_documento   = tipos.get((r.get("tipo_documento") or "").upper()),
                numero_documento = r.get("numero_documento") or None,
                primer_nombre    = r["primer_nombre"],
                segundo_nombre   = r.get("segundo_nombre") or None,
                primer_apellido  = r["primer_apellido"],
                segundo_apellido = r.get("segundo_apellido") or None,
                email            = r.get("email") or None,
                telefono         = r.get("telefono") or None,
                usuario_creacion = usuario_id,
            )
            creados += 1
        return creados


# ════════════════════════════════════════════════════════════
# 7 — POSTULACIONES (relacionar vacante por nombre,
#     candidato por tipo+num documento)
# ════════════════════════════════════════════════════════════

class ImportarPostulacionesView(APIView):
    parser_classes = [MultiPartParser]
    COLUMNAS_REQUERIDAS = ["vacante", "tipo_documento_candidato", "numero_documento_candidato"]

    def post(self, request, compania):
        return _procesar_importacion(
            request, self.COLUMNAS_REQUERIDAS, self._validar, self._importar,
            compania_id=compania)

    def _validar(self, rows, compania_id, **kw):
        from apps.vacantes.models import Vacante
        from apps.candidatos.models import TipoDocumento, Candidato, DatosCandidato
        log = []
        vacantes = {v.descripcion.upper(): v.id
                    for v in Vacante.objects.filter(compania=compania_id, ind_activa=True)}
        tipos_doc = {t.descripcion.upper(): t.id for t in TipoDocumento.objects.all()}
        for r in rows:
            f = r["_fila"]
            vac_nombre = r.get("vacante", "").upper()
            if not vac_nombre or vac_nombre not in vacantes:
                log.append(_err(f, f"Vacante '{r.get('vacante','?')}' no encontrada o inactiva")); continue
            tipo_d = r.get("tipo_documento_candidato","").upper()
            num_d  = r.get("numero_documento_candidato","").strip()
            if tipo_d not in tipos_doc:
                log.append(_err(f, f"tipo_documento_candidato '{tipo_d}' no existe")); continue
            tid = tipos_doc[tipo_d]
            if not DatosCandidato.objects.filter(
                compania=compania_id, tipo_documento=tid, numero_documento=num_d).exists():
                log.append(_err(f, f"Candidato con {tipo_d} {num_d} no encontrado")); continue
            log.append(_ok(f, f"Postulación '{r.get('vacante')}' ← cand {num_d} lista"))
        return log

    @transaction.atomic
    def _importar(self, rows, compania_id, usuario_id, **kw):
        from apps.vacantes.models import Vacante
        from apps.candidatos.models import (TipoDocumento, Candidato,
                                            DatosCandidato, EstadoPostulacion, Postulacion)
        vacantes   = {v.descripcion.upper(): v for v in Vacante.objects.filter(compania=compania_id, ind_activa=True)}
        tipos_doc  = {t.descripcion.upper(): t for t in TipoDocumento.objects.all()}
        est_rec    = EstadoPostulacion.objects.filter(descripcion="Recibida").first()
        n = Postulacion.objects.filter(compania=compania_id).count()
        creados = 0
        for r in rows:
            vac      = vacantes[r["vacante"].upper()]
            tid      = tipos_doc[r["tipo_documento_candidato"].upper()]
            datos_c  = DatosCandidato.objects.filter(
                compania=compania_id, tipo_documento=tid,
                numero_documento=r["numero_documento_candidato"].strip()).first()
            n += 1
            # Crear postulación — el trigger genera el intento + token automáticamente
            Postulacion.objects.create(
                compania            = _comp(compania_id),
                id_interno          = n,
                vacante             = vac,
                candidato           = datos_c.candidato,
                estado              = est_rec,
                descripcion         = r.get("descripcion") or None,
                fecha_postulacion   = timezone.now(),
                usuario_postulacion = usuario_id,
                usuario_creacion    = usuario_id,
            )
            # El trigger generará el token y el intento automáticamente.
            # El envío de email del token se realiza aquí:
            _enviar_email_evaluacion(datos_c, vac, compania_id)
            creados += 1
        return creados


def _enviar_email_evaluacion(datos_c, vac, compania_id):
    """Envía email con enlace de evaluación al candidato recién postulado."""
    from apps.candidatos.models import PostulacionToken
    if not datos_c.email:
        return
    try:
        token_obj = PostulacionToken.objects.filter(
            compania=compania_id,
            postulacion__candidato=datos_c.candidato,
            postulacion__vacante=vac,
        ).order_by("-fecha_creacion").first()
        if not token_obj:
            return
        fe = getattr(settings, "FRONTEND_URL", "http://localhost:5173")
        enlace = f"{fe}/evaluacion/acceso?token={token_obj.token}&llave={token_obj.llave}"
        nombre = f"{datos_c.primer_nombre} {datos_c.primer_apellido}"
        exp_fmt = token_obj.fecha_expiracion.strftime("%d/%m/%Y a las %H:%M")
        send_mail(
            f"Pander RRHH — Evaluación: {str(vac)[:60]}",
            f"Hola {nombre},\n\nHas sido postulado/a a:\n  {vac.descripcion}\n\n"
            f"Accede a tu evaluación:\n  {enlace}\n\n"
            f"Válido hasta: {exp_fmt}\n\nEquipo Pander RRHH",
            settings.DEFAULT_FROM_EMAIL, [datos_c.email], fail_silently=True,
        )
    except Exception:
        pass


# ════════════════════════════════════════════════════════════
# DESCARGA DE PLANTILLAS
# ════════════════════════════════════════════════════════════

class PlantillaExcelView(APIView):
    """GET /api/importacion/plantilla/<entidad>/"""

    PLANTILLAS = {
        "companias":     [["descripcion","nit","objeto_social","representante_legal","direccion","telefono","ind_activa","ind_evaluacion_vacante"],
                          ["Empresa Ejemplo","900123456-1","Comercio","Juan Pérez","Calle 10","601555","TRUE","FALSE"]],
        "analistas":     [["primer_nombre","segundo_nombre","primer_apellido","segundo_apellido","tipo_documento","numero_documento","telefono","cargo"],
                          ["María","Lucía","García","López","Cédula de Ciudadanía, Cédula de Extranjería, Pasaporte, NIT, Permiso por Protección Temporal","1020304050","3001234567","Analista de RRHH"]],
        "usuarios":      [["email","rol","tipo_documento_analista","numero_documento_analista","ind_super_usuario"],
                          ["analista@empresa.com","Analista","Cédula de Ciudadanía, Cédula de Extranjería, Pasaporte, NIT, Permiso por Protección Temporal","1020304050","FALSE"]],
        "unidades":      [["descripcion","especialidad"],
                          ["Tecnología","Desarrollo de Software"]],
        "vacantes":      [["descripcion","unidad","estado","tipo_contrato","anio_experiencia","salario_minimo","salario_maximo","ind_activa","ind_publicada"],
                          ["Desarrollador Backend","Tecnología","Abierta","Indefinido","2","3500000","5000000","TRUE","TRUE"]],
        "candidatos":    [["primer_nombre","segundo_nombre","primer_apellido","segundo_apellido","tipo_documento","numero_documento","email","telefono"],
                          ["Carlos","Andrés","Ramírez","Torres","Cédula de Ciudadanía, Cédula de Extranjería, Pasaporte, NIT, Permiso por Protección Temporal","1234567890","carlos@email.com","3209876543"]],
        "postulaciones": [["vacante","tipo_documento_candidato","numero_documento_candidato","descripcion"],
                          ["Desarrollador Backend","Cédula de Ciudadanía, Cédula de Extranjería, Pasaporte, NIT, Permiso por Protección Temporal","1234567890","Candidato referido"]],
    }

    def get(self, request, entidad):
        if entidad not in self.PLANTILLAS:
            return Response({"error": f"Entidad '{entidad}' no soportada."}, status=400)
        if openpyxl is None:
            return Response({"error": "openpyxl no instalado."}, status=500)

        from django.http import HttpResponse
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = entidad.capitalize()
        filas = self.PLANTILLAS[entidad]
        for row in filas:
            ws.append(row)
        # Estilo: cabeceras en negrita y color
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill("solid", fgColor="1E3A5F")
        for cell in ws[1]:
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col) + 4
            ws.column_dimensions[col[0].column_letter].width = max_len

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"] = f'attachment; filename="plantilla_{entidad}.xlsx"'
        return resp


# ════════════════════════════════════════════════════════════
# HELPER: procesar importación (2 fases: validar → importar)
# ════════════════════════════════════════════════════════════

def _comp(compania_id):
    from apps.empresa.models import Compania
    return Compania.objects.get(id=compania_id)


def _procesar_importacion(request, cols_req, validar_fn, importar_fn, compania_id=None):
    """
    Flujo estándar de importación:
      1. Lee el archivo Excel.
      2. Verifica que tenga las columnas requeridas.
      3. Ejecuta validar_fn → log de validación.
      4. Si dry_run=true → devuelve solo el log (sin importar).
      5. Si hay errores y no es dry_run → devuelve error 422.
      6. Si todo OK → importar_fn → devuelve resumen.
    """
    archivo = request.FILES.get("archivo")
    if not archivo:
        return Response({"error": "Se requiere el archivo Excel (campo 'archivo')."}, status=400)

    dry_run  = str(request.data.get("dry_run", "false")).lower() == "true"
    uid      = request.data.get("usuario_id", 1)
    try:
        uid = int(uid)
    except (ValueError, TypeError):
        uid = 1

    try:
        rows = _leer_excel(archivo)
    except Exception as e:
        return Response({"error": f"No se pudo leer el archivo: {e}"}, status=400)

    if not rows:
        return Response({"error": "El archivo está vacío o sin datos."}, status=400)

    # Verificar columnas requeridas
    if cols_req:
        cols_presentes = set(rows[0].keys()) - {"_fila"}
        faltantes = [c for c in cols_req if c not in cols_presentes]
        if faltantes:
            return Response({
                "error": f"Columnas requeridas faltantes: {', '.join(faltantes)}",
                "columnas_encontradas": sorted(cols_presentes),
            }, status=400)

    # Validar
    log = validar_fn(rows, compania_id=compania_id, usuario_id=uid)
    errores = [l for l in log if l["estado"] == "ERROR"]
    total_ok = len([l for l in log if l["estado"] == "OK"])

    if dry_run:
        return Response({
            "modo":     "validacion",
            "total":    len(rows),
            "ok":       total_ok,
            "errores":  len(errores),
            "log":      log,
        })

    if errores:
        return Response({
            "error":   "Hay errores de validación. Corrígelos antes de importar.",
            "errores": len(errores),
            "log":     log,
        }, status=422)

    # Importar
    try:
        creados = importar_fn(rows, compania_id=compania_id, usuario_id=uid)
    except Exception as e:
        return Response({"error": f"Error durante la importación: {e}"}, status=500)

    return Response({
        "modo":    "importacion",
        "creados": creados,
        "log":     log,
    }, status=201)
